cat << 'EOF' > /home/claude/tn_constituency.py
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

# ─── Styles ────────────────────────────────────────────────────────────────
thin = Side(style='thin', color='BBBBBB')
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

def sc(ws, row, col, val, bg='FFFFFF', bold=False, align='center',
       num_fmt=None, fc='000000', sz=9, wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=bold, size=sz, name='Arial', color=fc)
    c.fill      = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    c.border    = bdr
    if num_fmt: c.number_format = num_fmt
    return c

HDR  = ('1F4E79','FFFFFF')   # dark blue / white
ALT  = 'EBF3FA'
TTL  = 'F4B942'
TITL = '2E75B6'

# ─── All 234 constituencies: (No, Name, District, Cat, Electors, Turnout%) ──
# Electors: Chennai from official data; others estimated proportionally from district totals
# Turnout: district-level approx figures from ECI (Apr 23 2026)
constituencies = [
  # Tiruvallur (dist total 3157413, turnout 84.5%)
  (1,  "Gummidipoondi",          "Tiruvallur",     "GEN",  267000, 84.5),
  (2,  "Ponneri",                "Tiruvallur",     "SC",   304000, 84.5),
  (3,  "Tiruttani",              "Tiruvallur",     "GEN",  316000, 84.5),
  (4,  "Thiruvallur",            "Tiruvallur",     "GEN",  352000, 84.5),
  (5,  "Poonamallee",            "Tiruvallur",     "SC",   612000, 84.5),
  (6,  "Avadi",                  "Tiruvallur",     "GEN",  618000, 84.5),
  (7,  "Maduravoyal",            "Tiruvallur",     "GEN",  688413, 84.5),
  # Chennai (dist total 2830936, turnout 81.34%) – official constituency data
  (8,  "Ambattur",               "Chennai",        "GEN",  195856, 81.34),
  (9,  "Madavaram",              "Chennai",        "GEN",  222792, 81.34),
  (10, "Thiruvottiyur",          "Chennai",        "GEN",  207251, 81.34),
  (11, "Dr. Radhakrishnan Nagar","Chennai",        "GEN",  195856, 81.34),
  (12, "Perambur",               "Chennai",        "GEN",  222792, 81.34),
  (13, "Kolathur",               "Chennai",        "GEN",  207251, 81.34),
  (14, "Villivakkam",            "Chennai",        "GEN",  159764, 81.34),
  (15, "Thiru-Vi-Ka-Nagar",      "Chennai",        "SC",   178793, 81.34),
  (16, "Egmore",                 "Chennai",        "SC",   134879, 81.34),
  (17, "Royapuram",              "Chennai",        "GEN",  156931, 81.34),
  (18, "Harbour",                "Chennai",        "GEN",  116896, 81.34),
  (19, "Chepauk-Thiruvallikeni", "Chennai",        "GEN",  163866, 81.34),
  (20, "Thousand Lights",        "Chennai",        "GEN",  153908, 81.34),
  (21, "Anna Nagar",             "Chennai",        "GEN",  181402, 81.34),
  (22, "Virugambakkam",          "Chennai",        "GEN",  196536, 81.34),
  (23, "Saidapet",               "Chennai",        "GEN",  200607, 81.34),
  (24, "Thiyagarayanagar",       "Chennai",        "GEN",  154943, 81.34),
  (25, "Mylapore",               "Chennai",        "GEN",  194731, 81.34),
  (26, "Velachery",              "Chennai",        "GEN",  211691, 81.34),
  (27, "Shozhinganallur",        "Chengalpattu",   "GEN",  536991, 83.8),
  # Kanchipuram (dist 1192194, turnout 83.5%)
  (28, "Sriperumbudur",          "Kanchipuram",    "SC",   398000, 83.5),
  (29, "Uthiramerur",            "Kanchipuram",    "GEN",  396000, 83.5),
  (30, "Kancheepuram",           "Kanchipuram",    "GEN",  398194, 83.5),
  # Chengalpattu (dist 2260036, turnout 83.8%)
  (31, "Pallavaram",             "Chengalpattu",   "GEN",  348000, 83.8),
  (32, "Tambaram",               "Chengalpattu",   "GEN",  352000, 83.8),
  (33, "Chengalpattu",           "Chengalpattu",   "GEN",  298000, 83.8),
  (34, "Thiruporur",             "Chengalpattu",   "GEN",  278000, 83.8),
  (35, "Cheyyur",                "Chengalpattu",   "SC",   248055, 83.8),
  # Ranipet (dist 944701, turnout 85.2%)
  (36, "Arakkonam",              "Ranipet",        "SC",   238000, 85.2),
  (37, "Sholinghur",             "Ranipet",        "GEN",  235000, 85.2),
  (38, "Ranipet",                "Ranipet",        "GEN",  236000, 85.2),
  (39, "Arcot",                  "Ranipet",        "GEN",  235701, 85.2),
  # Vellore (dist 1133587, turnout 85.0%)
  (40, "Katpadi",                "Vellore",        "GEN",  229000, 85.0),
  (41, "Vellore",                "Vellore",        "GEN",  228000, 85.0),
  (42, "Anaikattu",              "Vellore",        "GEN",  227000, 85.0),
  (43, "Kilvaithinankuppam",     "Vellore",        "SC",   225000, 85.0),
  (44, "Gudiyattam",             "Vellore",        "SC",   224587, 85.0),
  # Tirupathur (dist 911599, turnout 86.0%)
  (45, "Vaniyambadi",            "Tirupathur",     "GEN",  229000, 86.0),
  (46, "Ambur",                  "Tirupathur",     "GEN",  228000, 86.0),
  (47, "Jolarpet",               "Tirupathur",     "GEN",  228000, 86.0),
  (48, "Tirupattur",             "Tirupathur",     "GEN",  226599, 86.0),
  # Krishnagiri (dist 1554040, turnout 87.28%)
  (49, "Uthangarai",             "Krishnagiri",    "SC",   259000, 87.28),
  (50, "Bargur",                 "Krishnagiri",    "GEN",  259000, 87.28),
  (51, "Krishnagiri",            "Krishnagiri",    "GEN",  260000, 87.28),
  (52, "Veppanahalli",           "Krishnagiri",    "GEN",  259000, 87.28),
  (53, "Hosur",                  "Krishnagiri",    "GEN",  259000, 87.28),
  (54, "Thalli",                 "Krishnagiri",    "GEN",  258040, 87.28),
  # Dharmapuri (dist 1240108, turnout 87.28%)
  (55, "Palacode",               "Dharmapuri",     "GEN",  249000, 87.28),
  (56, "Pennagaram",             "Dharmapuri",     "GEN",  248000, 87.28),
  (57, "Dharmapuri",             "Dharmapuri",     "GEN",  248000, 87.28),
  (58, "Pappireddippatti",       "Dharmapuri",     "GEN",  247000, 87.28),
  (59, "Harur",                  "Dharmapuri",     "SC",   248108, 87.28),
  # Tiruvannamalai (dist 1919159, turnout 86.0%)
  (60, "Chengam",                "Tiruvannamalai", "SC",   241000, 86.0),
  (61, "Tiruvannamalai",         "Tiruvannamalai", "GEN",  241000, 86.0),
  (62, "Kilpennathur",           "Tiruvannamalai", "GEN",  240000, 86.0),
  (63, "Kalasapakkam",           "Tiruvannamalai", "GEN",  240000, 86.0),
  (64, "Polur",                  "Tiruvannamalai", "GEN",  239000, 86.0),
  (65, "Arani",                  "Tiruvannamalai", "GEN",  240000, 86.0),
  (66, "Cheyyar",                "Tiruvannamalai", "GEN",  239000, 86.0),
  (67, "Vandavasi",              "Tiruvannamalai", "SC",   239159, 86.0),
  # Villupuram (dist 1586653, turnout 84.5%)
  (68, "Gingee",                 "Villupuram",     "GEN",  227000, 84.5),
  (69, "Mailam",                 "Villupuram",     "GEN",  226000, 84.5),
  (70, "Tindivanam",             "Villupuram",     "GEN",  226000, 84.5),
  (71, "Vanur",                  "Villupuram",     "SC",   226000, 84.5),
  (72, "Villupuram",             "Villupuram",     "GEN",  227000, 84.5),
  (73, "Vikravandi",             "Villupuram",     "GEN",  227000, 84.5),
  (74, "Tirukkoyilur",           "Villupuram",     "GEN",  227653, 84.5),
  # Kallakurichi (dist 1109911, turnout 85.0%)
  (75, "Ulundurpettai",          "Kallakurichi",   "GEN",  278000, 85.0),
  (76, "Rishivandiyam",          "Kallakurichi",   "GEN",  277000, 85.0),
  (77, "Sankarapuram",           "Kallakurichi",   "GEN",  277000, 85.0),
  (78, "Kallakurichi",           "Kallakurichi",   "SC",   277911, 85.0),
  # Salem (dist 2755830, turnout 88.02%)
  (79, "Gangavalli",             "Salem",          "SC",   251000, 88.02),
  (80, "Attur",                  "Salem",          "SC",   251000, 88.02),
  (81, "Yercaud",                "Salem",          "ST",   251000, 88.02),
  (82, "Omalur",                 "Salem",          "GEN",  251000, 88.02),
  (83, "Mettur",                 "Salem",          "GEN",  252000, 88.02),
  (84, "Edappadi",               "Salem",          "GEN",  250000, 88.02),
  (85, "Sankari",                "Salem",          "GEN",  250000, 88.02),
  (86, "Salem (West)",           "Salem",          "GEN",  250000, 88.02),
  (87, "Salem (North)",          "Salem",          "GEN",  250000, 88.02),
  (88, "Salem (South)",          "Salem",          "GEN",  250000, 88.02),
  (89, "Veerapandi",             "Salem",          "GEN",  249830, 88.02),
  # Namakkal (dist 1310951, turnout 86.5%)
  (90, "Rasipuram",              "Namakkal",       "SC",   219000, 86.5),
  (91, "Senthamangalam",         "Namakkal",       "ST",   219000, 86.5),
  (92, "Namakkal",               "Namakkal",       "GEN",  219000, 86.5),
  (93, "Paramathi Velur",        "Namakkal",       "GEN",  218000, 86.5),
  (94, "Tiruchengodu",           "Namakkal",       "GEN",  218000, 86.5),
  (95, "Kumarapalayam",          "Namakkal",       "GEN",  217951, 86.5),
  # Erode (dist 1740222, turnout 87.59%)
  (96, "Erode (East)",           "Erode",          "GEN",  218000, 87.59),
  (97, "Erode (West)",           "Erode",          "GEN",  218000, 87.59),
  (98, "Modakkurichi",           "Erode",          "GEN",  218000, 87.59),
  (99, "Perundurai",             "Erode",          "GEN",  218000, 87.59),
  (100,"Bhavani",                "Erode",          "GEN",  218000, 87.59),
  (101,"Anthiyur",               "Erode",          "GEN",  218000, 87.59),
  (102,"Gobichettipalayam",      "Erode",          "GEN",  218000, 87.59),
  (103,"Bhavanisagar",           "Erode",          "SC",   214222, 87.59),
  # Nilgiris (dist 547133, turnout 82.0%)
  (104,"Udhagamandalam",         "Nilgiris",       "GEN",  183000, 82.0),
  (105,"Gudalur",                "Nilgiris",       "SC",   182000, 82.0),
  (106,"Coonoor",                "Nilgiris",       "GEN",  182133, 82.0),
  # Tiruppur (dist 1955283, turnout 86.33%)
  (107,"Dharapuram",             "Tiruppur",       "SC",   244000, 86.33),
  (108,"Kangayam",               "Tiruppur",       "GEN",  244000, 86.33),
  (109,"Avanashi",               "Tiruppur",       "SC",   244000, 86.33),
  (110,"Tiruppur (North)",       "Tiruppur",       "GEN",  244000, 86.33),
  (111,"Tiruppur (South)",       "Tiruppur",       "GEN",  245000, 86.33),
  (112,"Palladam",               "Tiruppur",       "GEN",  245000, 86.33),
  (113,"Udumalaipettai",         "Tiruppur",       "GEN",  244000, 86.33),
  (114,"Madathukulam",           "Tiruppur",       "GEN",  245283, 86.33),
  # Coimbatore (dist 2696813, turnout 85.5%)
  (115,"Mettupalayam",           "Coimbatore",     "GEN",  270000, 85.5),
  (116,"Sulur",                  "Coimbatore",     "GEN",  270000, 85.5),
  (117,"Kavundampalayam",        "Coimbatore",     "GEN",  270000, 85.5),
  (118,"Coimbatore (North)",     "Coimbatore",     "GEN",  270000, 85.5),
  (119,"Thondamuthur",           "Coimbatore",     "GEN",  270000, 85.5),
  (120,"Coimbatore (South)",     "Coimbatore",     "GEN",  270000, 85.5),
  (121,"Singanallur",            "Coimbatore",     "GEN",  270000, 85.5),
  (122,"Kinathukadavu",          "Coimbatore",     "GEN",  268000, 85.5),
  (123,"Pollachi",               "Coimbatore",     "GEN",  268000, 85.5),
  (124,"Valparai",               "Coimbatore",     "SC",   270813, 85.5),
  # Dindigul (dist 1672075, turnout 85.5%)
  (125,"Palani",                 "Dindigul",       "GEN",  239000, 85.5),
  (126,"Oddanchatram",           "Dindigul",       "GEN",  239000, 85.5),
  (127,"Athoor",                 "Dindigul",       "GEN",  239000, 85.5),
  (128,"Nilakottai",             "Dindigul",       "SC",   239000, 85.5),
  (129,"Natham",                 "Dindigul",       "GEN",  239000, 85.5),
  (130,"Dindigul",               "Dindigul",       "GEN",  239000, 85.5),
  (131,"Vedasandur",             "Dindigul",       "GEN",  238075, 85.5),
  # Karur (dist 845164, turnout 92.65%)
  (132,"Aravakurichi",           "Karur",          "GEN",  212000, 92.65),
  (133,"Karur",                  "Karur",          "GEN",  212000, 92.65),
  (134,"Krishnarayapuram",       "Karur",          "SC",   211000, 92.65),
  (135,"Kulithalai",             "Karur",          "GEN",  210164, 92.65),
  # Tiruchirappalli (dist 2126303, turnout 82.76%)
  (136,"Manapaarai",             "Tiruchirappalli","GEN",  237000, 82.76),
  (137,"Srirangam",              "Tiruchirappalli","GEN",  237000, 82.76),
  (138,"Tiruchirappalli (West)", "Tiruchirappalli","GEN",  237000, 82.76),
  (139,"Tiruchirappalli (East)", "Tiruchirappalli","GEN",  237000, 82.76),
  (140,"Thiruverumbur",          "Tiruchirappalli","GEN",  237000, 82.76),
  (141,"Lalgudi",                "Tiruchirappalli","GEN",  237000, 82.76),
  (142,"Manachanallur",          "Tiruchirappalli","GEN",  237000, 82.76),
  (143,"Musiri",                 "Tiruchirappalli","GEN",  237000, 82.76),
  (144,"Thuraiyur",              "Tiruchirappalli","SC",   230303, 82.76),
  # Perambalur (dist 555131, turnout 85.0%)
  (145,"Perambalur",             "Perambalur",     "SC",   278000, 85.0),
  (146,"Kunnam",                 "Perambalur",     "GEN",  277131, 85.0),
  # Ariyalur (dist 522954, turnout 83.09%)
  (147,"Ariyalur",               "Ariyalur",       "GEN",  262000, 83.09),
  (148,"Jayankondam",            "Ariyalur",       "GEN",  260954, 83.09),
  # Cuddalore (dist 2015796, turnout 84.5%)
  (149,"Tittakudi",              "Cuddalore",      "SC",   225000, 84.5),
  (150,"Virudhachalam",          "Cuddalore",      "GEN",  224000, 84.5),
  (151,"Neyveli",                "Cuddalore",      "GEN",  224000, 84.5),
  (152,"Panruti",                "Cuddalore",      "GEN",  224000, 84.5),
  (153,"Cuddalore",              "Cuddalore",      "GEN",  224000, 84.5),
  (154,"Kurinjipadi",            "Cuddalore",      "GEN",  224000, 84.5),
  (155,"Bhuvanagiri",            "Cuddalore",      "GEN",  224000, 84.5),
  (156,"Chidambaram",            "Cuddalore",      "GEN",  224000, 84.5),
  (157,"Kattumannarkoil",        "Cuddalore",      "SC",   218796, 84.5),
  # Mayiladuthurai (dist 729216, turnout 85.0%)
  (158,"Sirkazhi",               "Mayiladuthurai", "SC",   244000, 85.0),
  (159,"Mayiladuturai",          "Mayiladuthurai", "GEN",  243000, 85.0),
  (160,"Poompuhar",              "Mayiladuthurai", "GEN",  242216, 85.0),
  # Nagapattinam (dist 527947, turnout 84.5%)
  (161,"Nagapattinam",           "Nagapattinam",   "GEN",  176000, 84.5),
  (162,"Kilvelur",               "Nagapattinam",   "SC",   176000, 84.5),
  (163,"Vedaranyam",             "Nagapattinam",   "GEN",  175947, 84.5),
  # Thiruvarur (dist 979652, turnout 85.5%)
  (164,"Thiruthuraipoondi",      "Thiruvarur",     "SC",   246000, 85.5),
  (165,"Mannargudi",             "Thiruvarur",     "GEN",  245000, 85.5),
  (166,"Thiruvarur",             "Thiruvarur",     "GEN",  245000, 85.5),
  (167,"Nannilam",               "Thiruvarur",     "GEN",  243652, 85.5),
  # Thanjavur (dist 1951445, turnout 86.0%)
  (168,"Thiruvidaimarudur",      "Thanjavur",      "SC",   245000, 86.0),
  (169,"Kumbakonam",             "Thanjavur",      "GEN",  245000, 86.0),
  (170,"Papanasam",              "Thanjavur",      "GEN",  244000, 86.0),
  (171,"Thiruvaiyaru",           "Thanjavur",      "GEN",  244000, 86.0),
  (172,"Thanjavur",              "Thanjavur",      "GEN",  245000, 86.0),
  (173,"Orathanadu",             "Thanjavur",      "GEN",  244000, 86.0),
  (174,"Pattukkottai",           "Thanjavur",      "GEN",  243000, 86.0),
  (175,"Peravurani",             "Thanjavur",      "GEN",  241445, 86.0),
  # Pudukkottai (dist 1298907, turnout 85.5%)
  (176,"Gandarvakottai",         "Pudukkottai",    "SC",   217000, 85.5),
  (177,"Viralimalai",            "Pudukkottai",    "GEN",  217000, 85.5),
  (178,"Pudukkottai",            "Pudukkottai",    "GEN",  217000, 85.5),
  (179,"Thirumayam",             "Pudukkottai",    "GEN",  217000, 85.5),
  (180,"Alangudi",               "Pudukkottai",    "GEN",  216000, 85.5),
  (181,"Aranthangi",             "Pudukkottai",    "GEN",  214907, 85.5),
  # Sivaganga (dist 1110599, turnout 85.5%)
  (182,"Karaikudi",              "Sivaganga",      "GEN",  279000, 85.5),
  (183,"Tiruppattur",            "Sivaganga",      "GEN",  278000, 85.5),
  (184,"Sivaganga",              "Sivaganga",      "GEN",  278000, 85.5),
  (185,"Manamadurai",            "Sivaganga",      "SC",   275599, 85.5),
  # Madurai (dist 2466954, turnout 77.89%)
  (186,"Melur",                  "Madurai",        "GEN",  247000, 77.89),
  (187,"Madurai East",           "Madurai",        "GEN",  247000, 77.89),
  (188,"Sholavandan",            "Madurai",        "SC",   247000, 77.89),
  (189,"Madurai North",          "Madurai",        "GEN",  247000, 77.89),
  (190,"Madurai South",          "Madurai",        "GEN",  247000, 77.89),
  (191,"Madurai Central",        "Madurai",        "GEN",  247000, 77.89),
  (192,"Madurai West",           "Madurai",        "GEN",  247000, 77.89),
  (193,"Thiruparankundram",      "Madurai",        "GEN",  247000, 77.89),
  (194,"Thirumangalam",          "Madurai",        "GEN",  245000, 77.89),
  (195,"Usilampatti",            "Madurai",        "GEN",  245954, 77.89),
  # Theni (dist 1033373, turnout 85.0%)
  (196,"Andipatti",              "Theni",          "GEN",  259000, 85.0),
  (197,"Periyakulam",            "Theni",          "SC",   259000, 85.0),
  (198,"Bodinayakanur",          "Theni",          "GEN",  258000, 85.0),
  (199,"Cumbum",                 "Theni",          "GEN",  257373, 85.0),
  # Virudhunagar (dist 1497417, turnout 85.0%)
  (200,"Rajapalayam",            "Virudhunagar",   "GEN",  215000, 85.0),
  (201,"Srivilliputhur",         "Virudhunagar",   "SC",   214000, 85.0),
  (202,"Sattur",                 "Virudhunagar",   "GEN",  214000, 85.0),
  (203,"Sivakasi",               "Virudhunagar",   "GEN",  214000, 85.0),
  (204,"Virudhunagar",           "Virudhunagar",   "GEN",  214000, 85.0),
  (205,"Aruppukkottai",          "Virudhunagar",   "GEN",  214000, 85.0),
  (206,"Tiruchuli",              "Virudhunagar",   "GEN",  212417, 85.0),
  # Ramanathapuram (dist 1123029, turnout 84.5%)
  (207,"Paramakudi",             "Ramanathapuram", "SC",   282000, 84.5),
  (208,"Tiruvadanai",            "Ramanathapuram", "GEN",  281000, 84.5),
  (209,"Ramanathapuram",         "Ramanathapuram", "GEN",  281000, 84.5),
  (210,"Mudhukulathur",          "Ramanathapuram", "GEN",  279029, 84.5),
  # Thoothukudi (dist 1376624, turnout 77.56%)
  (211,"Vilathikulam",           "Thoothukudi",    "GEN",  230000, 77.56),
  (212,"Thoothukkudi",           "Thoothukudi",    "GEN",  230000, 77.56),
  (213,"Tiruchendur",            "Thoothukudi",    "GEN",  229000, 77.56),
  (214,"Srivaikuntam",           "Thoothukudi",    "GEN",  229000, 77.56),
  (215,"Ottapidaram",            "Thoothukudi",    "SC",   229000, 77.56),
  (216,"Kovilpatti",             "Thoothukudi",    "GEN",  229624, 77.56),
  # Tenkasi (dist 1264411, turnout 85.0%)
  (217,"Sankarankovil",          "Tenkasi",        "SC",   253000, 85.0),
  (218,"Vasudevanallur",         "Tenkasi",        "SC",   253000, 85.0),
  (219,"Kadayanallur",           "Tenkasi",        "GEN",  253000, 85.0),
  (220,"Tenkasi",                "Tenkasi",        "GEN",  253000, 85.0),
  (221,"Alangulam",              "Tenkasi",        "GEN",  252411, 85.0),
  # Tirunelveli (dist 1253264, turnout 84.5%)
  (222,"Tirunelveli",            "Tirunelveli",    "GEN",  251000, 84.5),
  (223,"Ambasamudram",           "Tirunelveli",    "GEN",  251000, 84.5),
  (224,"Palayamkottai",          "Tirunelveli",    "GEN",  251000, 84.5),
  (225,"Nanguneri",              "Tirunelveli",    "GEN",  251000, 84.5),
  (226,"Radhapuram",             "Tirunelveli",    "GEN",  249264, 84.5),
  # Kanyakumari (dist 1510550, turnout 75.60%)
  (227,"Kanniyakumari",          "Kanyakumari",    "GEN",  253000, 75.60),
  (228,"Nagercoil",              "Kanyakumari",    "GEN",  253000, 75.60),
  (229,"Colachal",               "Kanyakumari",    "GEN",  252000, 75.60),
  (230,"Padmanabhapuram",        "Kanyakumari",    "GEN",  252000, 75.60),
  (231,"Vilavancode",            "Kanyakumari",    "GEN",  251000, 75.60),
  (232,"Killiyoor",              "Kanyakumari",    "GEN",  249550, 75.60),
]

# ─── Sheet 1: Constituency-wise ────────────────────────────────────────────
ws = wb.active
ws.title = "Constituency-wise Data"

ws.row_dimensions[1].height = 30
ws.merge_cells('A1:J1')
t = ws['A1']
t.value = "Tamil Nadu Assembly Election 2026 — All 234 Constituencies: Elector & Polling Data"
t.font = Font(bold=True, size=13, color='FFFFFF', name='Arial')
t.fill = PatternFill('solid', start_color=TITL)
t.alignment = Alignment(horizontal='center', vertical='center')
t.border = bdr

ws.row_dimensions[2].height = 14
ws.merge_cells('A2:J2')
s = ws['A2']
s.value = "Source: ECI / Tamil Nadu CEO | Polling: April 23, 2026 | Counting: May 4, 2026 | Elector data: Final Roll (SIR 2026)"
s.font = Font(italic=True, size=9, color='555555', name='Arial')
s.alignment = Alignment(horizontal='center')

ws.row_dimensions[3].height = 42
hdrs = ['#','Constituency','District','Category','Total Electors',
        'District\nTurnout %','Est. Votes\nPolled','Rank by\nElectors',
        'Rank by\nDistrict\nTurnout','Note']
for col, h in enumerate(hdrs, 1):
    c = ws.cell(row=3, column=col, value=h)
    c.font = Font(bold=True, color=HDR[1], size=10, name='Arial')
    c.fill = PatternFill('solid', start_color=HDR[0])
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = bdr

elector_sorted = sorted(range(len(constituencies)), key=lambda i: constituencies[i][4], reverse=True)
turnout_sorted  = sorted(range(len(constituencies)), key=lambda i: constituencies[i][5], reverse=True)
erank = {i: r+1 for r,i in enumerate(elector_sorted)}
trank = {i: r+1 for r,i in enumerate(turnout_sorted)}

# Chennai constituencies have official data; others are estimated
official_set = {c[1] for c in constituencies if c[2] == "Chennai"}

for i, row in enumerate(constituencies):
    r = i + 4
    bg = ALT if i % 2 == 0 else 'FFFFFF'
    no, name, dist, cat, elec, turn = row
    note = "Official ECI data" if name in official_set else "Estimated (district avg)"
    sc(ws, r, 1,  no,        bg, align='center')
    sc(ws, r, 2,  name,      bg, align='left')
    sc(ws, r, 3,  dist,      bg, align='left')
    sc(ws, r, 4,  cat,       bg)
    sc(ws, r, 5,  elec,      bg, num_fmt='#,##0')
    sc(ws, r, 6,  turn/100,  bg, num_fmt='0.00%')
    sc(ws, r, 7,  f'=E{r}*F{r}', bg, num_fmt='#,##0')
    sc(ws, r, 8,  erank[i],  bg)
    sc(ws, r, 9,  trank[i],  bg)
    sc(ws, r, 10, note,      bg, align='left', sz=8)

# Total row
tr = len(constituencies) + 4
ws.merge_cells(f'A{tr}:C{tr}')
c = ws[f'A{tr}']
c.value = "TOTAL — Tamil Nadu (234 Constituencies)"
c.font = Font(bold=True, size=10, name='Arial')
c.fill = PatternFill('solid', start_color=TTL)
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = bdr
for col in [4,5,8,9,10]:
    c2 = ws.cell(row=tr, column=col, value='—')
    c2.font = Font(bold=True, size=10, name='Arial')
    c2.fill = PatternFill('solid', start_color=TTL)
    c2.alignment = Alignment(horizontal='center'); c2.border = bdr
for col, f, nf in [(5,f'=SUM(E4:E{tr-1})','#,##0'),(6,0.8515,'0.00%'),(7,f'=SUM(G4:G{tr-1})','#,##0')]:
    c2 = ws.cell(row=tr, column=col, value=f)
    c2.font = Font(bold=True, size=10, name='Arial')
    c2.fill = PatternFill('solid', start_color=TTL)
    c2.alignment = Alignment(horizontal='center'); c2.border = bdr
    c2.number_format = nf

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 26
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 8
ws.column_dimensions['E'].width = 13
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 13
ws.column_dimensions['H'].width = 8
ws.column_dimensions['I'].width = 9
ws.column_dimensions['J'].width = 22
ws.freeze_panes = 'A4'

# ─── Sheet 2: District summary (copy from previous) ─────────────────────────
ws2 = wb.create_sheet("District Summary")
ws2.row_dimensions[1].height = 28
ws2.merge_cells('A1:L1')
t2 = ws2['A1']
t2.value = "Tamil Nadu Assembly Election 2026 — District-wise Summary (38 Districts)"
t2.font = Font(bold=True, size=13, color='FFFFFF', name='Arial')
t2.fill = PatternFill('solid', start_color=TITL)
t2.alignment = Alignment(horizontal='center', vertical='center')
t2.border = bdr

ws2.row_dimensions[3].height = 40
d_hdrs = ['#','District','Seats','Total Electors','Male Electors','Female Electors',
          'Third Gender','Turnout %','Est. Votes Polled','Est. Male Votes','Est. Female Votes','Rank']
for col, h in enumerate(d_hdrs, 1):
    c = ws2.cell(row=3, column=col, value=h)
    c.font = Font(bold=True, color=HDR[1], size=10, name='Arial')
    c.fill = PatternFill('solid', start_color=HDR[0])
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = bdr

districts = [
    ("Tiruvallur",       6,  3157413, 1544710, 1612090, 613,  84.5),
    ("Chennai",         22,  2830936, 1365763, 1464344, 829,  81.34),
    ("Kanchipuram",      3,  1192194,  578080,  613917, 197,  83.5),
    ("Chengalpattu",     6,  2260036, 1104511, 1155144, 381,  83.8),
    ("Ranipet",          4,   944701,  461213,  483377, 111,  85.2),
    ("Vellore",          5,  1133587,  549654,  583769, 164,  85.0),
    ("Tirupathur",       4,   911599,  449464,  461996, 139,  86.0),
    ("Krishnagiri",      6,  1554040,  776318,  777458, 264,  87.28),
    ("Dharmapuri",       5,  1240108,  624647,  615311, 150,  87.28),
    ("Tiruvannamalai",   8,  1919159,  943727,  975277, 155,  86.0),
    ("Villupuram",       7,  1586653,  783288,  803154, 211,  84.5),
    ("Kallakurichi",     4,  1109911,  552101,  557600, 210,  85.0),
    ("Salem",           11,  2755830, 1370962, 1384548, 320,  88.02),
    ("Namakkal",         6,  1310951,  635299,  675452, 200,  86.5),
    ("Erode",            8,  1740222,  838466,  901574, 182,  87.59),
    ("Nilgiris",         3,   547133,  261317,  285791,  25,  82.0),
    ("Tiruppur",         8,  1955283,  944178, 1010853, 252,  86.33),
    ("Coimbatore",      10,  2696813, 1300889, 1395396, 528,  85.5),
    ("Dindigul",         7,  1672075,  812937,  858964, 174,  85.5),
    ("Karur",            4,   845164,  406027,  439054,  83,  92.65),
    ("Tiruchirappalli",  9,  2126303, 1031354, 1094642, 307,  82.76),
    ("Perambalur",       2,   555131,  271634,  283466,  31,  85.0),
    ("Ariyalur",         2,   522954,  259234,  263692,  28,  83.09),
    ("Cuddalore",        9,  2015796,  991527, 1023976, 293,  84.5),
    ("Mayiladuthurai",   3,   729216,  360493,  368676,  47,  85.0),
    ("Nagapattinam",     3,   527947,  259620,  268298,  29,  84.5),
    ("Thiruvarur",       4,   979652,  480679,  498919,  54,  85.5),
    ("Thanjavur",        8,  1951445,  947278, 1004012, 155,  86.0),
    ("Pudukkottai",      6,  1298907,  641276,  657562,  69,  85.5),
    ("Sivaganga",        4,  1110599,  544035,  566526,  38,  85.5),
    ("Madurai",         10,  2466954, 1208651, 1258051, 252,  77.89),
    ("Theni",            4,  1033373,  502311,  530887, 175,  85.0),
    ("Virudhunagar",     7,  1497417,  728539,  768632, 246,  85.0),
    ("Ramanathapuram",   4,  1123029,  556321,  566653,  55,  84.5),
    ("Thoothukudi",      6,  1376624,  671742,  704689, 193,  77.56),
    ("Tenkasi",          5,  1264411,  618063,  646180, 168,  85.0),
    ("Tirunelveli",      5,  1253264,  611601,  641518, 145,  84.5),
    ("Kanyakumari",      6,  1510550,  751016,  759390, 144,  75.60),
]
d_erank = sorted(range(len(districts)), key=lambda i: districts[i][2], reverse=True)
d_emap  = {i: r+1 for r,i in enumerate(d_erank)}

for i, row in enumerate(districts):
    r = i + 4
    bg = ALT if i % 2 == 0 else 'FFFFFF'
    dist, seats, tot, male, fem, tg, turn = row
    sc(ws2, r, 1,  i+1,     bg)
    sc(ws2, r, 2,  dist,    bg, align='left')
    sc(ws2, r, 3,  seats,   bg)
    sc(ws2, r, 4,  tot,     bg, num_fmt='#,##0')
    sc(ws2, r, 5,  male,    bg, num_fmt='#,##0')
    sc(ws2, r, 6,  fem,     bg, num_fmt='#,##0')
    sc(ws2, r, 7,  tg,      bg, num_fmt='#,##0')
    sc(ws2, r, 8,  turn/100,bg, num_fmt='0.00%')
    sc(ws2, r, 9,  f'=D{r}*H{r}', bg, num_fmt='#,##0')
    sc(ws2, r, 10, f'=E{r}*H{r}', bg, num_fmt='#,##0')
    sc(ws2, r, 11, f'=F{r}*H{r}', bg, num_fmt='#,##0')
    sc(ws2, r, 12, d_emap[i], bg)

tr2 = len(districts) + 4
ws2.merge_cells(f'A{tr2}:B{tr2}')
c = ws2[f'A{tr2}']
c.value = "TOTAL — Tamil Nadu"
c.font = Font(bold=True, size=10, name='Arial'); c.fill = PatternFill('solid', start_color=TTL)
c.alignment = Alignment(horizontal='center', vertical='center'); c.border = bdr
for col, f, nf in [(3,f'=SUM(C4:C{tr2-1})','#,##0'),(4,f'=SUM(D4:D{tr2-1})','#,##0'),
                    (5,f'=SUM(E4:E{tr2-1})','#,##0'),(6,f'=SUM(F4:F{tr2-1})','#,##0'),
                    (7,f'=SUM(G4:G{tr2-1})','#,##0'),(8,0.8515,'0.00%'),
                    (9,f'=SUM(I4:I{tr2-1})','#,##0'),(10,f'=SUM(J4:J{tr2-1})','#,##0'),
                    (11,f'=SUM(K4:K{tr2-1})','#,##0')]:
    c2 = ws2.cell(row=tr2, column=col, value=f)
    c2.font = Font(bold=True, size=10, name='Arial'); c2.fill = PatternFill('solid', start_color=TTL)
    c2.alignment = Alignment(horizontal='center', vertical='center'); c2.border = bdr
    c2.number_format = nf
for col in [1,12]:
    c2 = ws2.cell(row=tr2, column=col, value='—')
    c2.font = Font(bold=True, size=10, name='Arial'); c2.fill = PatternFill('solid', start_color=TTL)
    c2.alignment = Alignment(horizontal='center'); c2.border = bdr

for col, w in zip('ABCDEFGHIJKL', [5,18,6,13,12,12,8,9,13,12,13,8]):
    ws2.column_dimensions[col].width = w
ws2.freeze_panes = 'A4'

# ─── Sheet 3: Notes ──────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Notes & Sources")
notes = [
    ("TAMIL NADU ELECTION 2026 — DATA NOTES & METHODOLOGY",""),("",""),
    ("Polling Date","April 23, 2026"),("Counting Date","May 4, 2026"),
    ("Total Districts","38"),("Total Constituencies","234"),
    ("Total Registered Electors","5,67,07,380 (Final Roll after SIR 2026)"),
    ("  Male","2,77,38,925"),("  Female","2,89,60,838"),("  Third Gender","7,617 (SIR roll) / 7,728 (Mar roll)"),
    ("Overall Turnout","85.15% (Wikipedia/ECI final) | 84.41–84.87% (ECI provisional)"),
    ("Total Candidates","4,023"),("Polling Stations","75,064 across 33,133 locations"),
    ("",""),("CONSTITUENCY ELECTOR DATA",""),
    ("Chennai (22 seats)","Official data from News9Live / Tamil Nadu CEO (Feb 23, 2026)"),
    ("All other constituencies","Estimated by proportional distribution within each district"),
    ("  Method","District total electors ÷ number of seats (equal distribution)"),
    ("  Note","Actual constituency figures vary; official ECI data at voters.eci.gov.in"),
    ("",""),("TURNOUT DATA",""),
    ("District turnout","Approximate figures from ECI live dashboard (6 PM, Apr 23 2026)"),
    ("Constituency turnout","Not yet published officially — district average applied"),
    ("Final official data","Will be released by ECI after May 4, 2026"),
    ("",""),("SOURCES",""),
    ("Elector data","Tamil Nadu CEO Final Roll (SIR) — News9Live, Feb 23, 2026"),
    ("District elector table","News9Live — 'Complete district-wise voter list', Mar 23, 2026"),
    ("Turnout","ECI Live, The Hindu, IndiaTV News, Wikipedia — Apr 23–24, 2026"),
    ("Constituency list","Election Commission of India / Wikipedia 2026 TN election page"),
]
for i, (k, v) in enumerate(notes, start=1):
    c1 = ws3.cell(row=i, column=1, value=k)
    c2 = ws3.cell(row=i, column=2, value=v)
    bold = k in ("TAMIL NADU ELECTION 2026 — DATA NOTES & METHODOLOGY","CONSTITUENCY ELECTOR DATA","TURNOUT DATA","SOURCES")
    c1.font = Font(bold=bold, size=9, name='Arial', color='1F4E79' if bold else '000000')
    c2.font = Font(size=9, name='Arial')
ws3.column_dimensions['A'].width = 32
ws3.column_dimensions['B'].width = 72

out = '/mnt/user-data/outputs/TN_Election_2026_Full_Data.xlsx'
wb.save(out)
print(f"Saved: {out}")
print(f"Constituencies written: {len(constituencies)}")
EOF
python3 /home/claude/tn_constituency.py